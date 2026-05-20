"""공용 xlsx 헬퍼 — 한 시트에 헤더 + 데이터 행을 쓰는 단순 빌더.

퀴즈/플래시카드 등 시트 한 장짜리 산출물에서 공통 호출. openpyxl 의존.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence


def write_table_xlsx(
    out_path: Path,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    sheet_name: str = "Sheet1",
) -> Path:
    """헤더 1행 + 데이터 행들로 채워진 .xlsx 파일을 저장하고 그 경로를 돌려준다.

    헤더는 굵게 + 옅은 음영. 컬럼 너비는 헤더·셀 길이의 max 로 자동 조정 (max 60).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel 시트명 31자 제한

    header_font = Font(bold=True, color="1F2A44")
    header_fill = PatternFill("solid", fgColor="E7EEFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    max_widths = [len(str(h)) for h in headers]
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap
            text = str(value) if value is not None else ""
            longest_line = max((len(line) for line in text.splitlines()), default=0)
            if col_idx <= len(max_widths):
                max_widths[col_idx - 1] = max(max_widths[col_idx - 1], longest_line)

    for col_idx, width in enumerate(max_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max(12, width + 2), 60
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
